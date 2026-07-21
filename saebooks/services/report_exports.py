"""Shared report-export serialization — CSV + XLSX.

Jurisdiction-neutral core. The two public serializers (:func:`build_csv`,
:func:`build_xlsx`) know nothing about accounting or tax: a caller hands them
a header row, a list of data rows, and the *column indices* that hold money.
Jurisdiction-specific exports (e.g. the AU BAS) build their own
``(headers, rows, money_cols)`` tuple in the jurisdiction module and reuse
these serializers, so the neutral core stays tax-free.

Money handling (the non-negotiable: never floats in the output)
--------------------------------------------------------------
The report pydantic schemas surface some amounts as ``float`` (a historical
API-boundary choice) and some as ``Decimal`` (e.g. the aged/cashbook models).
At the *export* boundary every money column is coerced back to an exact 2-dp
``Decimal`` via :func:`to_money` (``Decimal(str(x))`` round-trips the intended
decimal for any money value — which never has >15 significant digits — then
``quantize`` to cents):

* **CSV** — the Decimal is formatted ``f"{d:.2f}"``: two decimals, no thousands
  separator, no currency symbol (raw numbers, matching the pre-existing
  ``profit_loss.csv`` / ``trial_balance.csv`` convention).
* **XLSX** — the *Decimal itself* is written into the cell (openpyxl stores it
  as an exact numeric, not a binary float) with a ``#,##0.00`` number format so
  the spreadsheet displays two decimals with thousands separators.

Formula-injection guard (CWE-1236)
----------------------------------
Free-text columns are passed through :func:`csv_sanitize_text`, which prefixes a
leading apostrophe onto any value starting with a spreadsheet formula trigger
(``= + - @``). This matters for XLSX too: openpyxl treats a bare string that
starts with ``=`` as a *formula*, so sanitizing also stops a malicious contact
name from being written as a live formula cell.
"""
from __future__ import annotations

import csv as _csv
import io
from collections.abc import Sequence
from decimal import Decimal
from typing import Any

# Mirrors ``saebooks.services.reports._CSV_FORMULA_TRIGGER_CHARS`` (kept local so
# this neutral module has no dependency back into the report service layer).
_CSV_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")

_CENTS = Decimal("0.01")
_XLSX_MONEY_FORMAT = "#,##0.00"


def csv_sanitize_text(value: str) -> str:
    """Quote-prefix a free-text cell that starts with a formula trigger.

    Renders the value as literal text in Excel/Sheets/LibreOffice rather than
    letting it be evaluated as a formula. Identical behaviour to the existing
    ``reports._csv_sanitize_text`` guard.
    """
    if value and value[0] in _CSV_FORMULA_TRIGGER_CHARS:
        return f"'{value}"
    return value


def to_money(value: Any) -> Decimal:
    """Coerce a report amount (Decimal | float | int | str | None) to 2-dp Decimal.

    ``Decimal(str(x))`` avoids binary-float artefacts: a float ``1234.5`` becomes
    ``Decimal("1234.5")`` (via its ``str``), never ``Decimal(1234.5)``'s long
    binary tail. Quantized to cents (banker-safe ROUND_HALF_EVEN default).
    """
    if value is None:
        return Decimal("0.00")
    d = value if isinstance(value, Decimal) else Decimal(str(value))
    return d.quantize(_CENTS)


def _render_cell(value: Any, *, is_money: bool, is_text: bool) -> str:
    if is_money:
        return f"{to_money(value):.2f}"
    if value is None:
        return ""
    text = str(value)
    return csv_sanitize_text(text) if is_text else text


def build_csv(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    money_cols: Sequence[int] = (),
    text_cols: Sequence[int] = (),
) -> str:
    """Serialize a table to RFC 4180 CSV text.

    ``money_cols`` and ``text_cols`` are 0-based column indices. Money columns
    are formatted to two raw decimals; text columns get the formula-injection
    guard. Everything else is stringified verbatim.
    """
    money = set(money_cols)
    text = set(text_cols)
    buf = io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow(list(headers))
    for row in rows:
        writer.writerow(
            [
                _render_cell(cell, is_money=i in money, is_text=i in text)
                for i, cell in enumerate(row)
            ]
        )
    return buf.getvalue()


def build_xlsx(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    sheet_title: str = "Report",
    money_cols: Sequence[int] = (),
    text_cols: Sequence[int] = (),
    title_lines: Sequence[str] = (),
) -> bytes:
    """Serialize a table to an ``.xlsx`` workbook (single sheet) → bytes.

    ``title_lines`` (e.g. company name, report title, period) are written as a
    small header block above a blank spacer row and the bold column-header row;
    the first title line is emphasised. Money columns are written as exact
    Decimals with a ``#,##0.00`` number format.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    money = set(money_cols)
    text = set(text_cols)

    wb = Workbook()
    ws = wb.active
    # Excel caps sheet titles at 31 chars and forbids []:*?/\
    safe_title = "".join(c for c in sheet_title if c not in "[]:*?/\\")[:31] or "Report"
    ws.title = safe_title

    r = 1
    for idx, line in enumerate(title_lines):
        cell = ws.cell(row=r, column=1, value=line)
        cell.font = Font(bold=True, size=14) if idx == 0 else Font(italic=True)
        r += 1
    if title_lines:
        r += 1  # blank spacer row

    header_row = r
    for c, head in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=str(head))
        cell.font = Font(bold=True)
    r = header_row + 1

    for row in rows:
        for c, value in enumerate(row, start=1):
            idx = c - 1
            if idx in money:
                cell = ws.cell(row=r, column=c, value=to_money(value))
                cell.number_format = _XLSX_MONEY_FORMAT
            elif value is None:
                ws.cell(row=r, column=c, value="")
            elif idx in text:
                ws.cell(row=r, column=c, value=csv_sanitize_text(str(value)))
            elif isinstance(value, (int, float)):
                ws.cell(row=r, column=c, value=value)
            else:
                ws.cell(row=r, column=c, value=str(value))
        r += 1

    # Roughly size columns to the widest cell (headers + data), capped.
    for c in range(1, len(headers) + 1):
        letter = get_column_letter(c)
        width = len(str(headers[c - 1]))
        for row in rows:
            if c - 1 < len(row) and row[c - 1] is not None:
                width = max(width, len(str(row[c - 1])))
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 48)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
