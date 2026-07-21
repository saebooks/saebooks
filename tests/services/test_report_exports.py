"""Unit tests for the neutral report-export serializers.

No DB — pure serialization. Verifies the money-never-float guarantee (the XLSX
XML stores exact 2-dp decimals, not binary-float artefacts) and the
formula-injection guard for both CSV and XLSX.
"""
from __future__ import annotations

import io
import re
import zipfile
from decimal import Decimal

import openpyxl

from saebooks.services.report_exports import build_csv, build_xlsx, to_money


def test_to_money_round_trips_float_without_binary_artifact() -> None:
    # 0.1 + 0.2 == 0.30000000000000004 as a float; to_money must yield 0.30.
    assert to_money(0.1 + 0.2) == Decimal("0.30")
    assert to_money(1234.5) == Decimal("1234.50")
    assert to_money(Decimal("99.995")) == Decimal("100.00")  # ROUND_HALF_EVEN
    assert to_money(None) == Decimal("0.00")


def test_build_csv_money_and_formula_guard() -> None:
    headers = ["name", "amount"]
    rows = [["=cmd()", 1234.5], ["Rent", Decimal("99.99")]]
    text = build_csv(headers, rows, money_cols=[1], text_cols=[0])
    lines = text.splitlines()
    assert lines[0] == "name,amount"
    # leading '=' is neutralised with an apostrophe; money is raw 2-dp
    assert lines[1] == "'=cmd(),1234.50"
    assert lines[2] == "Rent,99.99"


def test_build_xlsx_stores_exact_decimals() -> None:
    rows = [["a", 0.1 + 0.2], ["b", Decimal("1234.56")]]
    content = build_xlsx(["k", "amt"], rows, money_cols=[1])
    assert content[:2] == b"PK"
    z = zipfile.ZipFile(io.BytesIO(content))
    xml = z.read("xl/worksheets/sheet1.xml").decode()
    values = re.findall(r"<v>([^<]+)</v>", xml)
    # exact stored numerics — no 0.30000000000000004
    assert "0.3" in values
    assert "1234.56" in values
    assert not any("0000000" in v for v in values)


def test_build_xlsx_money_number_format_and_guard() -> None:
    rows = [["=EVIL", Decimal("5.00")]]
    content = build_xlsx(["name", "amt"], rows, money_cols=[1], text_cols=[0], title_lines=["Co", "Report", "2026"])
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    # locate header row
    header_row = None
    for r in ws.iter_rows():
        if r[0].value == "name":
            header_row = r[0].row
            break
    assert header_row is not None
    name_cell = ws.cell(row=header_row + 1, column=1)
    amt_cell = ws.cell(row=header_row + 1, column=2)
    assert name_cell.value == "'=EVIL"  # formula neutralised
    assert amt_cell.number_format == "#,##0.00"
